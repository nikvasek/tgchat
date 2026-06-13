from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .messages import HEADERS, FoundMessage


class ExcelExporter:
    def __init__(self, output_file: str):
        self._path = Path(output_file)
        self._known_ids: dict[str, set[str]] = {}

    def _get_workbook(self) -> Workbook:
        if self._path.exists():
            return load_workbook(self._path)
        return Workbook()

    def _get_worksheet(self, workbook: Workbook, title: str) -> Worksheet:
        if title in workbook.sheetnames:
            return workbook[title]

        if len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet":
            sheet = workbook.active
            sheet.title = title
            sheet.append(HEADERS)
            return sheet

        sheet = workbook.create_sheet(title=title)
        sheet.append(HEADERS)
        return sheet

    def _load_existing_ids(self, sheet: Worksheet, sheet_title: str) -> set[str]:
        if sheet_title in self._known_ids:
            return self._known_ids[sheet_title]

        ids: set[str] = set()
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) > 1:
            header = [str(cell) if cell is not None else "" for cell in rows[0]]
            try:
                id_index = header.index("ID сообщения")
                source_index = header.index("Источник")
                chat_index = header.index("Username чата")
                title_index = header.index("Чат")
            except ValueError:
                id_index = len(HEADERS) - 1
                source_index = 1
                chat_index = 3
                title_index = 2

            for row in rows[1:]:
                if not row or len(row) <= id_index or row[id_index] is None:
                    continue
                source = str(row[source_index]) if len(row) > source_index and row[source_index] else ""
                chat = str(row[chat_index]) if len(row) > chat_index and row[chat_index] else ""
                if not chat and len(row) > title_index and row[title_index]:
                    chat = str(row[title_index])
                ids.add(f"{source}:{chat}:{row[id_index]}")

        self._known_ids[sheet_title] = ids
        return ids

    def append_messages(self, sheet_title: str, messages: list[FoundMessage]) -> int:
        if not messages:
            return 0

        workbook = self._get_workbook()
        sheet = self._get_worksheet(workbook, sheet_title)
        existing = self._load_existing_ids(sheet, sheet_title)

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        added = 0

        for message in messages:
            if message.dedup_key in existing:
                continue
            sheet.append(message.to_row(now))
            existing.add(message.dedup_key)
            added += 1

        if workbook.sheetnames and workbook.sheetnames[0] == "Sheet" and len(workbook.sheetnames) > 1:
            del workbook["Sheet"]

        self._path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(self._path)
        self._known_ids[sheet_title] = existing
        return added
